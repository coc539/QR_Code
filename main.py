import qrcode
import os
import json
import threading
import customtkinter as ctk
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from tkinter import messagebox, Toplevel
from datetime import datetime

class QRCodeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Code Generator")
        self.root.geometry("400x500")
        self.root.minsize(400, 500)

        # Set initial default labels
        self.default_labels = {"A": "Data A", "B": "Data B", "C": "Data C", "D": "Data D", "E": "Data E"}
        self.labels_path = "labels_config.json"
        self.excel_path_config = "excel_path_config.json"
        self.entry_labels = self.load_labels()

        # Excel and Image directories setup
        self.excel_directory = "excel_files"
        os.makedirs(self.excel_directory, exist_ok=True)
        self.image_directory = "qr_code_images"
        os.makedirs(self.image_directory, exist_ok=True)

        # Initialize the current Excel file path
        self.main_excel_path = self.load_or_get_excel_path()
        self.initialize_excel_file()

        # Track previous labels to determine when to create a new Excel file
        self.previous_labels = self.entry_labels.copy()

        # Frame for entry fields
        self.entry_frame = ctk.CTkFrame(root)
        self.entry_frame.pack(pady=20)

        # Entry fields and labels in the main window
        self.entries = {}
        self.labels = {}
        row = 0
        for key in self.entry_labels:
            label = ctk.CTkLabel(self.entry_frame, text=self.entry_labels[key], font=("Arial", 12))
            label.grid(row=row, column=0, padx=10, pady=5, sticky="w")
            self.labels[key] = label

            entry = ctk.CTkEntry(self.entry_frame, width=250, font=("Arial", 12))
            entry.grid(row=row, column=1, padx=10, pady=5)
            self.entries[key] = entry
            row += 1

        # Button to generate QR Code
        self.generate_button = ctk.CTkButton(root, text="Generate QR Code", command=self.generate_qr, font=("Arial", 12), fg_color="green", text_color="white")
        self.generate_button.pack(pady=20)

        # Label to display the QR Code
        self.qr_code_label = ctk.CTkLabel(root, text="")
        self.qr_code_label.pack()

        # Bind Ctrl+Q to open the label edit window
        self.root.bind('<Control-q>', self.open_label_edit_window)
        self.root.after(1000, self.ensure_ctrl_q_binding)  # Ensures binding remains active

    def load_labels(self):
        if os.path.exists(self.labels_path):
            with open(self.labels_path, 'r') as file:
                return json.load(file)
        return self.default_labels.copy()

    def save_labels(self, labels):
        with open(self.labels_path, 'w') as file:
            json.dump(labels, file)

    def load_or_get_excel_path(self):
        date_str = datetime.now().strftime("%Y%m%d")
        daily_directory = os.path.join(self.excel_directory, date_str)
        os.makedirs(daily_directory, exist_ok=True)

        if os.path.exists(self.excel_path_config):
            with open(self.excel_path_config, 'r') as file:
                excel_path = json.load(file).get("path")
                if excel_path and os.path.exists(excel_path):
                    return excel_path

        excel_path = os.path.join(daily_directory, f'qr_codes_{date_str}.xlsx')
        with open(self.excel_path_config, 'w') as file:
            json.dump({"path": excel_path}, file)
        return excel_path

    def initialize_excel_file(self):
        if not os.path.exists(self.main_excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "QR Codes"
            sheet.append(list(self.entry_labels.values()) + ["QR Code Image"])
            for cell in sheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                sheet.column_dimensions[cell.column_letter].width = 20
            sheet.column_dimensions['F'].width = 25
            workbook.save(self.main_excel_path)

    def open_label_edit_window(self, event=None):
        # Check if the edit window already exists and is still open
        if hasattr(self, 'edit_window') and self.edit_window.winfo_exists():
            self.edit_window.lift()  # Bring existing window to the front
            return

        # Create a new label edit window
        self.edit_window = Toplevel(self.root)
        self.edit_window.title("Edit Labels")
        self.edit_window.geometry("500x400")

        # Use a frame for label entry fields
        self.edit_frame = ctk.CTkFrame(self.edit_window)
        self.edit_frame.pack(pady=20, padx=20)

        self.label_entries = {}
        row = 0
        for key, label_text in self.entry_labels.items():
            label = ctk.CTkLabel(self.edit_frame, text=f"Edit Label for {label_text}:", font=("Arial", 12))
            label.grid(row=row, column=0, padx=10, pady=5, sticky="w")
            label_entry = ctk.CTkEntry(self.edit_frame, width=200, font=("Arial", 12))  # Entry box starts empty
            label_entry.grid(row=row, column=1, padx=10, pady=5)
            self.label_entries[key] = label_entry
            row += 1

        submit_button = ctk.CTkButton(self.edit_window, text="Submit", command=self.submit_labels, font=("Arial", 12), fg_color="blue", text_color="white")
        submit_button.pack(pady=20)

    def submit_labels(self):
        new_labels = {}
        for key in self.entry_labels:
            entry_value = self.label_entries[key].get().strip()
            new_labels[key] = entry_value if entry_value else self.entry_labels[key]

        if new_labels != self.previous_labels:
            self.entry_labels = new_labels
            self.save_labels(self.entry_labels)

            # Update labels in the main window
            for key, label in self.labels.items():
                label.configure(text=self.entry_labels[key])

            date_str = datetime.now().strftime("%Y%m%d")
            daily_directory = os.path.join(self.excel_directory, date_str)
            os.makedirs(daily_directory, exist_ok=True)
            self.main_excel_path = os.path.join(daily_directory, f'qr_codes_{date_str}_{datetime.now().strftime("%H%M%S")}.xlsx')

            self.initialize_excel_file()

            with open(self.excel_path_config, 'w') as file:
                json.dump({"path": self.main_excel_path}, file)

            self.previous_labels = new_labels.copy()

        self.edit_window.destroy()

    def generate_qr(self):
        # Run QR generation in a separate thread to keep UI responsive
        threading.Thread(target=self._generate_qr_task).start()

    def _generate_qr_task(self):
        data_values = [self.entries[key].get().strip() for key in ['A', 'B', 'C', 'D', 'E']]
        filtered_data_values = [value for value in data_values if value]

        if not filtered_data_values:
            messagebox.showwarning("Warning", "At least one field must contain data.")
            return

        self.save_to_excel(self.main_excel_path, list(self.entry_labels.values()), data_values)

        for entry in self.entries.values():
            entry.delete(0, 'end')

    def save_to_excel(self, excel_path, headers, data_values):
        filtered_data = [value for value in data_values if value]
        data = ",".join(filtered_data)

        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=4)
        qr.add_data(data)
        qr.make(fit=True)

        qr_image = qr.make_image(fill="black", back_color="white")
        image_filename = f"{self.image_directory}/{data.replace(' ', '_').replace(',', '_')}.png"
        qr_image.save(image_filename)

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        row_number = sheet.max_row + 1
        sheet.append(data_values)

        for col, _ in enumerate(data_values, start=1):
            cell = sheet.cell(row=row_number, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.row_dimensions[row_number].height = 100
        img = ExcelImage(image_filename)
        img.width, img.height = 90, 90
        sheet.add_image(img, f"F{row_number}")

        try:
            workbook.save(excel_path)
            messagebox.showinfo("Success", f"QR Code and data saved to '{excel_path}'")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")

        qr_image_pil = Image.open(image_filename)
        qr_photo = ctk.CTkImage(qr_image_pil, size=(200, 200))
        self.qr_code_label.configure(image=qr_photo)
        self.qr_code_label.image = qr_photo

    def ensure_ctrl_q_binding(self):
        self.root.bind('<Control-q>', self.open_label_edit_window)
        self.root.after(1000, self.ensure_ctrl_q_binding)

# Run the application                    
if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    app = QRCodeApp(root)
    root.mainloop()
